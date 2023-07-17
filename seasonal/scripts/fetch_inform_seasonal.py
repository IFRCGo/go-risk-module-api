import openpyxl

from seasonal.models import InformRiskSeasonal
from common.models import HazardType, Country


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_hazard_type(hazard_type):
    if hazard_type == "INFORM.TC.Seasonal":
        hazard_type = HazardType.CYCLONE
    elif hazard_type == "INFORM.DR.Seasonal":
        hazard_type = HazardType.DROUGHT
    elif hazard_type == "INFORM.FL.Seasonal":
        hazard_type = HazardType.FLOOD
    return hazard_type


def fetch_inform_seasonal(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name("Sheet1")
    max_rows = get_maximum_rows(sheet_object=worksheet)
    hazard_list = [HazardType.CYCLONE, HazardType.DROUGHT, HazardType.FLOOD]
    for i in range(2, max_rows + 1):
        iso3 = worksheet.cell(row=i, column=1).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=2).value)
        january = worksheet.cell(row=i, column=3).value
        february = worksheet.cell(row=i, column=4).value
        march = worksheet.cell(row=i, column=5).value
        april = worksheet.cell(row=i, column=6).value
        may = worksheet.cell(row=i, column=7).value
        june = worksheet.cell(row=i, column=8).value
        july = worksheet.cell(row=i, column=9).value
        august = worksheet.cell(row=i, column=10).value
        september = worksheet.cell(row=i, column=11).value
        october = worksheet.cell(row=i, column=12).value
        november = worksheet.cell(row=i, column=13).value
        december = worksheet.cell(row=i, column=14).value

        if Country.objects.filter(iso3=iso3.lower()).exists() and hazard_type in hazard_list:
            data = {
                "country": Country.objects.filter(
                    iso3=iso3.lower(),
                    independent=True,
                    is_deprecated=False,
                ).first(),
                "hazard_type": hazard_type,
                "january": january,
                "february": february,
                "march": march,
                "april": april,
                "may": may,
                "june": june,
                "july": july,
                "august": august,
                "september": september,
                "october": october,
                "november": november,
                "december": december,
            }
            InformRiskSeasonal.objects.create(**data)
