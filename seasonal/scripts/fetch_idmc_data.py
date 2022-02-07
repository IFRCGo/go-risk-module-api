import openpyxl

from common.models import HazardType, Country
from seasonal.models import Idmc


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_hazard_type(hazard_type):
    if hazard_type == 'flood':
        hazard_type = HazardType.FLOOD
    elif hazard_type == 'storm':
        hazard_type = HazardType.STORM
    elif hazard_type == 'food_insecurity':
        hazard_type = HazardType.FOOD_INSECURITY
    return hazard_type


def fetch_idmc_data(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name('Displacements Risk per month')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    # Iterate the loop to read the cell values
    for i in range(2, max_rows + 1):
        country = worksheet.cell(row=i, column=1).value
        iso3 = worksheet.cell(row=i, column=2).value
        hazard_type = worksheet.cell(row=i, column=3).value.lower()
        annual_average_displacement = worksheet.cell(row=i, column=4).value
        if annual_average_displacement == 'No data':
            annual_average_displacement = None
        elif annual_average_displacement != 'No data':
            annual_average_displacement = float(str(annual_average_displacement).replace(',', ''))
        confidence_type = worksheet.cell(row=i, column=5).value.lower()
        if confidence_type in ['Less than 10 data points in the 39 years covered by the dataset', 'No data']:
            confidence_type = 'undefined'
        note = worksheet.cell(row=i, column=6).value
        january = worksheet.cell(row=i, column=7).value
        if january == '':
            january = None
        february = worksheet.cell(row=i, column=8).value
        if february == '':
            february = None
        march = worksheet.cell(row=i, column=9).value
        if march == '':
            march = None
        april = worksheet.cell(row=i, column=10).value
        if april == '':
            april = None
        may = worksheet.cell(row=i, column=11).value
        if may == '':
            may = None
        june = worksheet.cell(row=i, column=12).value
        if june == '':
            june = None
        july = worksheet.cell(row=i, column=13).value
        if july == '':
            july = None
        august = worksheet.cell(row=i, column=14).value
        if august == '':
            august = None
        september = worksheet.cell(row=i, column=15).value
        if september == '':
            september = None
        october = worksheet.cell(row=i, column=16).value
        if october == '':
            october = None
        november = worksheet.cell(row=i, column=17).value
        if november == '':
            november = None
        december = worksheet.cell(row=i, column=18).value
        if december == '':
            december = None

        data = {
            'country': Country.objects.filter(iso3=iso3.lower()).first(),
            'iso3': iso3,
            'hazard_type': parse_hazard_type(hazard_type),
            'annual_average_displacement': annual_average_displacement,
            'confidence_type': confidence_type,
            'note': note,
            'january': january,
            'february': february,
            'march': march,
            'april': april,
            'may': may,
            'june': june,
            'july': july,
            'august': august,
            'september': september,
            'october': october,
            'november': november,
            'december': december
        }
        Idmc.objects.create(**data)
