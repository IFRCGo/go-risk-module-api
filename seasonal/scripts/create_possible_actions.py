import openpyxl
import datetime

from common.models import HazardType, Country
from seasonal.models import (
    PossibleEarlyActions,
    PossibleEarlyActionsSectors,
)


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def get_merge_lookup(sheet, cell):
    """
    Used to extract cell value from merged_cells
    """
    idx = cell.coordinate
    for range_ in sheet.merged_cells.ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
        for row in merged_cells:
            if idx in row:
                return sheet[merged_cells[0][0]].value
    return sheet[idx].value


def parse_date(date):
    try:
        date = datetime.datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        return None


def parse_budget_cost(budget):
    if budget:
        budget_split = budget.split(" ")
        return int(budget_split[1].replace(",", ""))
    return None


def parse_number_of_people_at_risk(people):
    if people and isinstance(people, str):
        people_split = people.split(" ")
        if len(people_split) > 1:
            new_people_at_risk = float(people_split[0]) * 1000000
            return new_people_at_risk
    return people


def parse_number_of_people_covered(people):
    try:
        if people:
            return int(people)
    except ValueError:
        return None


def parse_hazard_type(hazard_type):
    hazard_type_dict = {
        "Flood": HazardType.FLOOD,
        "Cyclone and Typhoon": HazardType.WIND,
        "Tropical Storm": HazardType.STORM,
        "Drought": HazardType.DROUGHT,
    }
    return hazard_type_dict.get(hazard_type)


def parse_country(country):
    if country:
        new_country = country.split(",")
        return new_country


def parse_exist_in_hub(value):
    if value == "NO":
        return False
    elif value == "YES":
        return True


def parse_evidence_of_sucess(text):
    if text:
        return text.replace("\\", " ")


def create_possible_actions(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name("Possible early actions")
    max_rows = get_maximum_rows(sheet_object=worksheet)
    # Iterate through loop to read the cell values
    for i in range(3, max_rows + 1):
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=1).value)
        country_list = worksheet.cell(row=i, column=4).value
        if country_list:
            countries = parse_country(country_list)
        early_actions = worksheet.cell(row=i, column=2).value
        hazard_name = worksheet.cell(row=i, column=3).value
        location = worksheet.cell(row=i, column=5).value
        sector = worksheet.cell(row=i, column=6).value
        intended_purpose = worksheet.cell(row=i, column=7).value
        organization = worksheet.cell(row=i, column=8).value
        budget = parse_budget_cost(worksheet.cell(row=i, column=9).value)
        cost = parse_budget_cost(worksheet.cell(row=i, column=10).value)
        implementation_date_raw = worksheet.cell(row=i, column=11).value
        timeframe_raw = worksheet.cell(row=i, column=12).value
        effective_time_raw = worksheet.cell(row=i, column=13).value
        number_of_people_covered = parse_number_of_people_covered(worksheet.cell(row=i, column=14).value)
        number_of_people_at_risk = parse_number_of_people_at_risk(worksheet.cell(row=i, column=15).value)
        scalability = worksheet.cell(row=i, column=16).value
        cross_cutting = worksheet.cell(row=i, column=17).value
        resources_used = worksheet.cell(row=i, column=18).value
        impact_action = worksheet.cell(row=i, column=19).value
        evidence_of_sucess = parse_evidence_of_sucess(worksheet.cell(row=i, column=20).value)
        resource = get_merge_lookup(worksheet, worksheet.cell(row=i, column=21))
        link_to_resources = get_merge_lookup(worksheet, worksheet.cell(row=i, column=22))
        exist_in_hub = parse_exist_in_hub(worksheet.cell(row=i, column=23).value)
        sector_list = sector.split(",")
        sector_append_list = []
        for sec in sector_list:
            sector = PossibleEarlyActionsSectors.objects.create(
                name=sec.strip().lower(),
            )
            sector_append_list.append(sector)

        for country in countries:
            if country:
                if Country.objects.filter(name__icontains=country).exists() and hazard_type:
                    data = {
                        "country": Country.objects.filter(name__icontains=country).first(),
                        "hazard_type": hazard_type,
                        "early_actions": early_actions,
                        "hazard_name": hazard_name,
                        "location": location,
                        "intended_purpose": intended_purpose,
                        "organization": organization,
                        "budget": budget,
                        "cost": cost,
                        "number_of_people_covered": number_of_people_covered,
                        "number_of_people_at_risk": number_of_people_at_risk,
                        "scalability": scalability,
                        "cross_cutting": cross_cutting,
                        "resources_used": resources_used,
                        "impact_action": impact_action,
                        "evidence_of_sucess": evidence_of_sucess,
                        "resource": resource,
                        "link_to_resources": link_to_resources,
                        "exist_in_hub": exist_in_hub,
                        "implementation_date_raw": implementation_date_raw,
                        "timeframe_raw": timeframe_raw,
                        "effective_time_raw": effective_time_raw,
                    }
                    possible_actions = PossibleEarlyActions.objects.create(**data)
                    possible_actions.sectors.add(*sector_append_list)
