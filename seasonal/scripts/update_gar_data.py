import openpyxl

from seasonal.models import GarHazardDisplacement
from common.models import HazardType, Country

from django.db import models


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_empty_cell_value(data):
    if data == "":
        data = {}
    return data


def parse_hazard_type(hazard_type):
    if hazard_type == "Flood":
        hazard_type = HazardType.FLOOD
    return hazard_type


def update_gar_data(file):
    # NOTE: Import File: seasonal/gar_update.xlsx
    # NOTE: set `data_only=True` to read cell value instead of formula
    workbook = openpyxl.load_workbook(file, data_only=True)
    worksheet = workbook.get_sheet_by_name("Economic loss")
    max_rows = get_maximum_rows(sheet_object=worksheet)
    for i in range(2, max_rows + 1):
        country = worksheet.cell(row=i, column=2).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=3).value)
        economic_loss_return_period_10_years = parse_empty_cell_value(worksheet.cell(row=i, column=4).value)
        economic_loss_return_period_25_years = parse_empty_cell_value(worksheet.cell(row=i, column=5).value)
        economic_loss_return_period_50_years = parse_empty_cell_value(worksheet.cell(row=i, column=6).value)
        economic_loss_return_period_100_years = parse_empty_cell_value(worksheet.cell(row=i, column=7).value)
        economic_loss_return_period_250_years = parse_empty_cell_value(worksheet.cell(row=i, column=8).value)
        economic_loss_return_period_500_years = parse_empty_cell_value(worksheet.cell(row=i, column=9).value)
        if Country.objects.filter(name=country).exists():
            if economic_loss_return_period_10_years:
                economic_loss_return_period_10_years = economic_loss_return_period_10_years * 1000000
            if economic_loss_return_period_25_years:
                economic_loss_return_period_25_years = economic_loss_return_period_25_years * 1000000
            if economic_loss_return_period_50_years:
                economic_loss_return_period_50_years = economic_loss_return_period_50_years * 1000000
            if economic_loss_return_period_100_years:
                economic_loss_return_period_100_years = economic_loss_return_period_100_years * 1000000
            if economic_loss_return_period_250_years:
                economic_loss_return_period_250_years = economic_loss_return_period_250_years * 1000000
            if economic_loss_return_period_500_years:
                economic_loss_return_period_500_years = economic_loss_return_period_500_years * 1000000
            gar_hazard = GarHazardDisplacement.objects.filter(
                country__name=country,
                hazard_type=hazard_type,
            )
            if gar_hazard.exists():
                continue
                """gar_hazard = gar_hazard.update(
                    ten_years=models.Func(
                        models.F('ten_years'),
                        models.Value(['economic_loss']),
                        models.Value(economic_loss_return_period_10_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    twenty_five_years=models.Func(
                        models.F('twenty_five_years'),
                        models.Value(['economic_loss']),
                        models.Value(economic_loss_return_period_25_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    fifty_years=models.Func(
                        models.F('fifty_years'),
                        models.Value(['economic_loss']),
                        models.Value(economic_loss_return_period_50_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    hundred_years=models.Func(
                        models.F('hundred_years'),
                        models.Value(['economic_loss']),
                        models.Value(economic_loss_return_period_100_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    two_hundred_fifty_years=models.Func(
                        models.F('two_hundred_fifty_years'),
                        models.Value(['economic_loss']),
                        models.Value(economic_loss_return_period_250_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    five_hundred_years=models.Func(
                        models.F('five_hundred_years'),
                        models.Value(['economic_loss']),
                        models.Value(economic_loss_return_period_500_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                )"""
            else:
                economic_loss_return_period_10_years = {"economic_loss": economic_loss_return_period_10_years}
                economic_loss_return_period_25_years = {"economic_loss": economic_loss_return_period_25_years}
                economic_loss_return_period_50_years = {"economic_loss": economic_loss_return_period_50_years}
                economic_loss_return_period_100_years = {"economic_loss": economic_loss_return_period_100_years}
                economic_loss_return_period_250_years = {"economic_loss": economic_loss_return_period_250_years}
                economic_loss_return_period_500_years = {"economic_loss": economic_loss_return_period_500_years}
                data = {
                    "country": Country.objects.filter(name=country).first(),
                    "hazard_type": hazard_type,
                    "ten_years": economic_loss_return_period_10_years,
                    "twenty_five_years": economic_loss_return_period_25_years,
                    "fifty_years": economic_loss_return_period_50_years,
                    "hundred_years": economic_loss_return_period_100_years,
                    "two_hundred_fifty_years": economic_loss_return_period_250_years,
                    "five_hundred_years": economic_loss_return_period_500_years,
                }
                GarHazardDisplacement.objects.get_or_create(**data)
    worksheet1 = workbook.get_sheet_by_name("Population exposure")
    max_rows = get_maximum_rows(sheet_object=worksheet1)
    for i in range(1, max_rows + 1):
        country = worksheet1.cell(row=i, column=2).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=3).value)
        population_exposure_return_period_10_years = parse_empty_cell_value(worksheet1.cell(row=i, column=4).value)
        population_exposure_return_period_25_years = parse_empty_cell_value(worksheet1.cell(row=i, column=5).value)
        population_exposure_return_period_50_years = parse_empty_cell_value(worksheet1.cell(row=i, column=6).value)
        if Country.objects.filter(name=country).exists():
            data = GarHazardDisplacement.objects.filter(
                country__name=country,
                hazard_type=hazard_type,
            ).update(
                ten_years=models.Func(
                    models.F("ten_years"),
                    models.Value(["population_exposure"]),
                    models.Value(population_exposure_return_period_10_years, models.JSONField()),
                    True,
                    function="jsonb_set",
                ),
                twenty_five_years=models.Func(
                    models.F("twenty_five_years"),
                    models.Value(["population_exposure"]),
                    models.Value(population_exposure_return_period_25_years, models.JSONField()),
                    True,
                    function="jsonb_set",
                ),
                fifty_years=models.Func(
                    models.F("fifty_years"),
                    models.Value(["population_exposure"]),
                    models.Value(population_exposure_return_period_50_years, models.JSONField()),
                    True,
                    function="jsonb_set",
                ),
            )
