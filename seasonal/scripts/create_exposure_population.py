import openpyxl

from oddrin.models import GarHazardDisplacement, HazardType


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_empty_cell_value(data):
    if data == '':
        data = None
    return data


def parse_hazard_type(hazard_type):
    if hazard_type == 'Flood':
        hazard_type = HazardType.FLOOD
    return hazard_type


def create_exposure_population(file):
    from ipc.models import Country

    # NOTE: set `data_only=True` to read cell value instead of formula
    workbook = openpyxl.load_workbook(file, data_only=True)
    worksheet = workbook.get_sheet_by_name('Flood')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    for i in range(1, max_rows):
        country = worksheet.cell(row=i, column=1).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=3).value)
        return_period_25_years = parse_empty_cell_value(worksheet.cell(row=i, column=4).value)
        return_period_50_years = parse_empty_cell_value(worksheet.cell(row=i, column=5).value)
        if Country.objects.filter(name=country).exists():
            data_flood = {
                'country': Country.objects.filter(name=country).first(),
                'hazard_type': hazard_type,
                'population_exposure_return_period_25_years': return_period_25_years,
                'population_exposure_return_period_50_years': return_period_50_years,
            }
            GarHazardDisplacement.objects.get_or_create(**data_flood)
