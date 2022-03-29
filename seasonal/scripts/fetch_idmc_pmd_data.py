import openpyxl

from django.db import models

from common.models import Country, HazardType
from seasonal.models import GarHazardDisplacement
from django.db.models.expressions import Func


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_hazard_type(hazard_type):
    if hazard_type == 'Earthquake':
        hazard_type = HazardType.EARTHQUAKE
    elif hazard_type == 'Storm-surge':
        hazard_type = HazardType.STORM
    elif hazard_type == 'Tsunami':
        hazard_type = HazardType.TSUNAMI
    elif hazard_type == 'Flood':
        hazard_type = HazardType.FLOOD
    elif hazard_type == 'Cyclonic Wind':
        hazard_type = HazardType.WIND
    return hazard_type


def parse_empty_cell_value(data):
    if data == '':
        data = None
    return data


class ReplaceValue(Func):

    function = 'jsonb_update'
    template = "%(function)s(%(expressions)s, '{\"%(keyname)s\"}','\"%(new_value)s\"', %(create_missing)s)"
    arity = 1

    def __init__(
        self, expression: str, keyname: str, new_value: int,
        create_missing: bool = False, **extra,
    ):
        super().__init__(
            expression,
            keyname=keyname,
            new_value=new_value,
            create_missing='true' if create_missing else 'false',
            **extra,
        )


def fetch_idmc_pmd_data(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name('ADD-PMD')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    for i in range(2, max_rows):
        country = worksheet.cell(row=i, column=2).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=3).value)
        population_displacement_return_period_20_years = parse_empty_cell_value(worksheet.cell(row=i, column=5).value)
        population_displacement_return_period_50_years = parse_empty_cell_value(worksheet.cell(row=i, column=6).value)
        population_displacement_return_period_100_years = parse_empty_cell_value(worksheet.cell(row=i, column=7).value)
        population_displacement_return_period_250_years = parse_empty_cell_value(worksheet.cell(row=i, column=8).value)
        population_displacement_return_period_1000_years = parse_empty_cell_value(worksheet.cell(row=i, column=9).value)
        population_displacement_return_period_1500_years = parse_empty_cell_value(worksheet.cell(row=i, column=10).value)
        if Country.objects.filter(name=country).exists():
            gar_hazard = GarHazardDisplacement.objects.filter(
                country__name=country,
                hazard_type=hazard_type,
            )
            if gar_hazard.exists():
                gar_hazard = gar_hazard.update(
                    twenty_years=models.Func(
                        models.F('twenty_years'),
                        models.Value(['population_displacement']),
                        models.Value(population_displacement_return_period_20_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    fifty_years=models.Func(
                        models.F('fifty_years'),
                        models.Value(['population_displacement']),
                        models.Value(population_displacement_return_period_50_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    hundred_years=models.Func(
                        models.F('hundred_years'),
                        models.Value(['population_displacement']),
                        models.Value(population_displacement_return_period_100_years, models.JSONField()),
                        True,
                        function="jsonb_set"
                    ),
                    two_hundred_fifty_years=models.Func(
                        models.F('two_hundred_fifty_years'),
                        models.Value(['population_displacement']),
                        models.Value(population_displacement_return_period_250_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    one_thousand_years=models.Func(
                        models.F('one_thousand_years'),
                        models.Value(['population_displacement']),
                        models.Value(population_displacement_return_period_1000_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                    one_thousand_five_hundred_years=models.Func(
                        models.F('one_thousand_five_hundred_years'),
                        models.Value(['population_displacement']),
                        models.Value(population_displacement_return_period_1500_years, models.JSONField()),
                        True,
                        function="jsonb_set",
                    ),
                )
            else:
                population_displacement_return_period_20_years = {
                    'population_displacement': population_displacement_return_period_20_years
                }
                population_displacement_return_period_50_years = {
                    'population_displacement': population_displacement_return_period_50_years
                }
                population_displacement_return_period_100_years = {
                    'population_displacement': population_displacement_return_period_100_years
                }
                population_displacement_return_period_250_years = {
                    'population_displacement': population_displacement_return_period_250_years
                }
                population_displacement_return_period_1000_years = {
                    'population_displacement': population_displacement_return_period_1000_years
                }
                population_displacement_return_period_1500_years = {
                    'population_displacement': population_displacement_return_period_1500_years
                }
                data = {
                    'country': Country.objects.get(name=country),
                    'hazard_type': hazard_type,
                    'twenty_years': population_displacement_return_period_20_years,
                    'fifty_years': population_displacement_return_period_50_years,
                    'hundred_years': population_displacement_return_period_100_years,
                    'two_hundred_fifty_years': population_displacement_return_period_250_years,
                    'one_thousand_years': population_displacement_return_period_1000_years,
                    'one_thousand_five_hundred_years': population_displacement_return_period_1500_years
                }
                GarHazardDisplacement.objects.get_or_create(**data)
