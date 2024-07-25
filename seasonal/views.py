from datetime import datetime

import pandas as pd
from django.db import models
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, extend_schema_view
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import response, status, viewsets
from rest_framework.decorators import action

from common.models import HazardType
from imminent.models import GWIS
from seasonal.filter_set import (
    PossibleEarlyActionsFilterSet,
    PublishReportFilterSet,
    RiskScoreFilterSet,
)
from seasonal.models import (
    DisplacementData,
    GarHazardDisplacement,
    GlobalDisplacement,
    GwisSeasonal,
    Idmc,
    IdmcSuddenOnset,
    InformRisk,
    InformRiskSeasonal,
    PossibleEarlyActions,
    PossibleEarlyActionsSectors,
    PublishReport,
    RiskScore,
    ThinkHazardInformation,
)
from seasonal.serializers import (
    CharKeyValueSerializer,
    DisplacementDataSerializer,
    GarHazardDisplacementSerializer,
    GlobalDisplacementSerializer,
    IdmcSerializer,
    IdmcSuddenOnsetSerializer,
    InformRiskSeasonalSerializer,
    InformRiskSerializer,
    InformScoreSerializer,
    PossibleEarlyActionOptionsSerializer,
    PossibleEarlyActionsSerializer,
    PublishReportSerializer,
    RiskScoreSerializer,
    SeasonalCountryRequestSerializer,
    SeasonalCountrySerializer,
    SeasonalRequestSerializer,
    SeasonalSerializer,
    ThinkHazardInformationSerializer,
)


class IdmcViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Idmc.objects.all()
    serializer_class = IdmcSerializer


class InformRiskViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = InformRisk.objects.select_related("country")
    serializer_class = InformRiskSerializer


class IdmcSuddenOnsetViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = IdmcSuddenOnset.objects.select_related("country")
    serializer_class = IdmcSuddenOnsetSerializer


class InformRiskSeasonalViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = InformRiskSeasonal.objects.select_related("country")
    serializer_class = InformRiskSeasonalSerializer


class DisplacementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DisplacementData.objects.select_related("country")
    serializer_class = DisplacementDataSerializer


class GarHazardViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GarHazardDisplacement.objects.select_related("country")
    serializer_class = GarHazardDisplacementSerializer


class GlobalDisplacementViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = GlobalDisplacementSerializer
    queryset = GlobalDisplacement.objects.all()


class ThinkHazardInformationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ThinkHazardInformationSerializer
    queryset = ThinkHazardInformation.objects.all()


@extend_schema_view(
    list=extend_schema(parameters=[SeasonalRequestSerializer], responses=SeasonalSerializer(many=True)),
)
class SeasonalViewSet(viewsets.ViewSet):
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ("iso3", "hazard_type")

    def list(self, request, *args, **kwargs):
        iso3 = self.request.query_params.get("iso3")
        region = self.request.query_params.get("region")

        idmc_qs = Idmc.objects.select_related("country")
        ipc_displacement_data_qs = GlobalDisplacement.objects.select_related("country")
        raster_displacement_data_qs = DisplacementData.objects.select_related("country")
        gwis_seasonal_qs = GwisSeasonal.objects.select_related("country")

        # XXX: Is this filters required? We have other enpoints for Country
        if iso3 is not None:
            idmc_qs = idmc_qs.filter(iso3__icontains=iso3)
            ipc_displacement_data_qs = ipc_displacement_data_qs.filter(country__iso3__icontains=iso3)
            raster_displacement_data_qs = raster_displacement_data_qs.filter(country__iso3__icontains=iso3)
            gwis_seasonal_qs = gwis_seasonal_qs.filter(country__iso3__icontains=iso3)
        # XXX: Is this filters used?
        elif region:
            idmc_qs = idmc_qs.filter(country__region__name=region)
            ipc_displacement_data_qs = ipc_displacement_data_qs.filter(country__region=region)
            raster_displacement_data_qs = raster_displacement_data_qs.filter(country__region__name=region)
            gwis_seasonal_qs = gwis_seasonal_qs.filter(country__region__name=region)

        data = {
            "idmc": idmc_qs,
            "ipc_displacement_data": ipc_displacement_data_qs,
            "raster_displacement_data": raster_displacement_data_qs,
            "gwis_seasonal": gwis_seasonal_qs,
        }
        return response.Response([SeasonalSerializer(data).data])


@extend_schema_view(
    list=extend_schema(parameters=[SeasonalCountryRequestSerializer], responses=SeasonalCountrySerializer(many=True)),
)
class SeasonalCountryViewSet(viewsets.ViewSet):

    # XXX: Unbound request
    def list(self, request, *args, **kwargs):
        iso3 = request.query_params.get("iso3")
        if iso3 is None:
            return response.Response(
                {"message": "Please provide iso3 in the query params"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        idmc_qs = Idmc.objects.select_related("country").filter(country__iso3__icontains=iso3)
        ipc_displacement_data_qs = GlobalDisplacement.objects.select_related("country").filter(country__iso3__icontains=iso3)
        raster_displacement_data_qs = DisplacementData.objects.select_related("country").filter(country__iso3__icontains=iso3)
        inform_qs = InformRisk.objects.select_related("country").filter(country__iso3__icontains=iso3)
        inform_seasonal_qs = InformRiskSeasonal.objects.select_related("country").filter(country__iso3__icontains=iso3)
        return_period_data_qs = GarHazardDisplacement.objects.select_related("country").filter(country__iso3__icontains=iso3)
        gwis_qs = GWIS.objects.select_related("country").filter(country__iso3__icontains=iso3)
        gwis_seasonal_qs = GwisSeasonal.objects.select_related("country").filter(country__iso3__icontains=iso3)

        data = {
            "idmc": idmc_qs,
            "ipc_displacement_data": ipc_displacement_data_qs,
            "raster_displacement_data": raster_displacement_data_qs,
            "inform": inform_qs,
            "inform_seasonal": inform_seasonal_qs,
            "return_period_data": return_period_data_qs,
            "gwis": gwis_qs,
            "gwis_seasonal": gwis_seasonal_qs,
            # 'idmc_return_period': idmc_return_period_data,
            # 'hazard_info': hazard_info,
            # 'gar_loss': gar_loss,
        }

        return response.Response([SeasonalCountrySerializer(data).data])


def generate_data(request):
    exposure_queryset = DisplacementData.objects.filter(
        models.Q(hazard_type=HazardType.FLOOD) | models.Q(hazard_type=HazardType.CYCLONE)
    )
    ipc_displacement = GlobalDisplacement.objects.filter(hazard_type=HazardType.FOOD_INSECURITY)
    all_data = []
    for exposure in exposure_queryset:
        data = dict(
            country=exposure.country.name,
            hazard_type=exposure.hazard_type,
            exposure_january=exposure.january,
            exposure_february=exposure.february,
            exposure_march=exposure.march,
            exposure_april=exposure.april,
            exposure_may=exposure.may,
            exposure_june=exposure.june,
            exposure_july=exposure.july,
            exposure_august=exposure.august,
            exposure_september=exposure.september,
            exposure_october=exposure.october,
            exposure_november=exposure.november,
            exposure_december=exposure.december,
            year="",
        )
        all_data.append(data)
    for ipc_data in ipc_displacement:
        new_data = dict(
            country=ipc_data.country.name,
            hazard_type=ipc_data.hazard_type,
            exposure_january=ipc_data.total_displacement if ipc_data.month == 1 else "",
            exposure_february=ipc_data.total_displacement if ipc_data.month == 2 else "",
            exposure_march=ipc_data.total_displacement if ipc_data.month == 3 else "",
            exposure_april=ipc_data.total_displacement if ipc_data.month == 4 else "",
            exposure_may=ipc_data.total_displacement if ipc_data.month == 5 else "",
            exposure_june=ipc_data.total_displacement if ipc_data.month == 6 else "",
            exposure_july=ipc_data.total_displacement if ipc_data.month == 7 else "",
            exposure_august=ipc_data.total_displacement if ipc_data.month == 8 else "",
            exposure_september=ipc_data.total_displacement if ipc_data.month == 9 else "",
            exposure_october=ipc_data.total_displacement if ipc_data.month == 10 else "",
            exposure_november=ipc_data.total_displacement if ipc_data.month == 11 else "",
            exposure_december=ipc_data.total_displacement if ipc_data.month == 12 else "",
            year=ipc_data.year,
        )
        all_data.append(new_data)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename={date}-exposure-summary.xlsx".format(
        date=datetime.now().strftime("%Y-%m-%d"),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Exposure Data"

    header_font = Font(name="Calibri", bold=True)

    columns = [
        ("country", 20),
        ("hazard_type", 20),
        ("exposure_january", 20),
        ("exposure_february", 20),
        ("exposure_march", 20),
        ("exposure_april", 20),
        ("exposure_may", 20),
        ("exposure_june", 20),
        ("exposure_july", 20),
        ("exposure_august", 20),
        ("exposure_september", 20),
        ("exposure_october", 20),
        ("exposure_november", 20),
        ("exposure_december", 20),
        ("year", 20),
    ]

    row_num = 1

    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    for event in all_data:
        row_num += 1
        row = [
            event["country"],
            event["hazard_type"],
            event["exposure_january"],
            event["exposure_february"],
            event["exposure_march"],
            event["exposure_april"],
            event["exposure_may"],
            event["exposure_june"],
            event["exposure_july"],
            event["exposure_august"],
            event["exposure_september"],
            event["exposure_october"],
            event["exposure_november"],
            event["exposure_december"],
            event["year"],
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response


class EarlyActionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PossibleEarlyActions.objects.select_related("country")
    filterset_class = PossibleEarlyActionsFilterSet
    serializer_class = PossibleEarlyActionsSerializer

    @extend_schema(request=None, responses=PossibleEarlyActionOptionsSerializer)
    @action(detail=False, url_path="options")
    def get_options(self, request, **kwargs):
        return response.Response(
            PossibleEarlyActionOptionsSerializer(
                dict(
                    sectors=PossibleEarlyActionsSectors.objects.all().distinct("name"),
                    hazard_type=CharKeyValueSerializer.choices_to_data(HazardType.choices),
                )
            ).data
        )


class PublishReportViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PublishReport.objects.all()
    filterset_class = PublishReportFilterSet
    serializer_class = PublishReportSerializer


class RiskScoreViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = RiskScore.objects.filter(country__isnull=False).select_related("country")
    serializer_class = RiskScoreSerializer
    filterset_class = RiskScoreFilterSet


class InformScoreViewSet(viewsets.ViewSet):
    @extend_schema(request=None, responses=InformScoreSerializer)
    def list(self, request, *args, **kwargs):
        def map_hazard_type(hazard_type):
            if hazard_type == "DR":
                return HazardType.DROUGHT
            elif hazard_type == "FL":
                return HazardType.FLOOD
            elif hazard_type == "TC":
                return HazardType.CYCLONE
            elif hazard_type == "FI":
                hazard_type = HazardType.FOOD_INSECURITY
            elif hazard_type == "EQ":
                hazard_type = HazardType.EARTHQUAKE
            elif hazard_type == "CD":
                hazard_type = HazardType.WIND
            elif hazard_type == "TS":
                hazard_type = HazardType.TSUNAMI
            elif hazard_type == "SS":
                hazard_type = HazardType.STORM
            return

        hazard_type = request.query_params.getlist("hazard_type")
        hazard_type_list = [token for line in hazard_type for token in line.split(",")]
        new_hazard_list = []
        for hazard_type in hazard_type_list:
            new_hazard_list.append(map_hazard_type(hazard_type))
        region = request.query_params.get("region")
        inform_risk_score_queryset = InformRiskSeasonal.objects.all()
        if region:
            inform_risk_score_queryset = inform_risk_score_queryset.filter(
                country__region=region,
            )
        if hazard_type:
            inform_risk_score_queryset = inform_risk_score_queryset.filter(
                hazard_type__in=new_hazard_list,
            )

        inform_score_dataframe = pd.DataFrame(
            list(
                inform_risk_score_queryset.values(
                    "country__name",
                    "country__iso3",
                    "hazard_type",
                    "january",
                    "february",
                    "march",
                    "april",
                    "may",
                    "june",
                    "july",
                    "august",
                    "september",
                    "october",
                    "november",
                    "december",
                )
            )
        )

        RISK_SCORE_CONSTANT = 10
        inform_score_dataframe.rename({"country__iso3": "iso3", "country__name": "name"}, axis=1, inplace=True)
        frame = pd.DataFrame([])
        for x, y in inform_score_dataframe.groupby(["iso3"]):
            for col in [
                "january",
                "february",
                "march",
                "april",
                "may",
                "june",
                "july",
                "august",
                "september",
                "october",
                "november",
                "december",
            ]:
                y[col] = y[col].sum()
            frame = pd.concat([frame, y], axis=0, ignore_index=True)
        frame = frame.rename({"hazard_type_x": "hazard_type", "name_x": "name"})
        frame["january"] = frame["january"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["february"] = frame["february"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["march"] = frame["march"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["april"] = frame["april"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["may"] = frame["may"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["june"] = frame["june"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["july"] = frame["july"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["august"] = frame["august"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["september"] = frame["september"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["october"] = frame["october"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["november"] = frame["november"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["december"] = frame["december"].apply(
            lambda x: x * RISK_SCORE_CONSTANT / frame.select_dtypes(exclude=["object"]).unstack().max()
        )
        frame["hazard_type"] = ",".join(list(frame["hazard_type"].unique()))
        frame = frame.drop_duplicates(["name"], keep="first")
        data = [
            {
                "country": {
                    "name": row["name"],
                    "iso3": row["iso3"],
                },
                "hazard_type": row["hazard_type"].split(","),
                "january": row["january"],
                "february": row["february"],
                "march": row["march"],
                "april": row["april"],
                "may": row["may"],
                "june": row["june"],
                "july": row["july"],
                "august": row["august"],
                "september": row["september"],
                "october": row["october"],
                "november": row["november"],
                "december": row["december"],
            }
            for index, row in frame.iterrows()
        ]
        return response.Response(InformScoreSerializer(data, many=True).data)
