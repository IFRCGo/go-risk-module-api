import openpyxl

from oddrin.models import GarHazard, HazardType


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows



def parse_empty_cell_value(data):
    if data == '':
        data = None
    return data


def create_gar_data(file):
    from ipc.models import Country

    # NOTE: set `data_only=True` to read cell value instead of formula
    workbook = openpyxl.load_workbook(file, data_only=True)
    worksheet = workbook.get_sheet_by_name('Sheet1')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    for i in range(3, max_rows):
        country = worksheet.cell(row=i, column=1).value
        wd_return_period_20_years = parse_empty_cell_value(worksheet.cell(row=i, column=6).value)
        wd_return_period_50_years = parse_empty_cell_value(worksheet.cell(row=i, column=7).value)
        wd_return_period_100_years = parse_empty_cell_value(worksheet.cell(row=i, column=8).value)
        wd_return_period_250_years = parse_empty_cell_value(worksheet.cell(row=i, column=9).value)
        wd_return_period_500_years = parse_empty_cell_value(worksheet.cell(row=i, column=10).value)
        ss_return_period_20_years = parse_empty_cell_value(worksheet.cell(row=i, column=11).value)
        ss_return_period_50_years = parse_empty_cell_value(worksheet.cell(row=i, column=12).value)
        ss_return_period_100_years = parse_empty_cell_value(worksheet.cell(row=i, column=13).value)
        ss_return_period_250_years = parse_empty_cell_value(worksheet.cell(row=i, column=14).value)
        ss_return_period_500_years = parse_empty_cell_value(worksheet.cell(row=i, column=15).value)
        if Country.objects.filter(name=country).exists():
            data_storm = {
                'country': Country.objects.filter(name=country).first(),
                'hazard_type': HazardType.STORM,
                'return_period_20_years': ss_return_period_20_years,
                'return_period_50_years': ss_return_period_50_years,
                'return_period_100_years': ss_return_period_100_years,
                'return_period_250_years': ss_return_period_250_years,
                'return_period_500_years': ss_return_period_500_years
            }
            GarHazard.objects.create(**data_storm)
            data_wind = {
                'country': Country.objects.filter(name=country).first(),
                'hazard_type': HazardType.WIND,
                'return_period_20_years': wd_return_period_20_years,
                'return_period_50_years': wd_return_period_50_years,
                'return_period_100_years': wd_return_period_100_years,
                'return_period_250_years': wd_return_period_250_years,
                'return_period_500_years': wd_return_period_500_years
            }
            GarHazard.objects.create(**data_wind)
