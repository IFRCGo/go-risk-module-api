import openpyxl

from oddrin.models import Oddrin, InformRisk, HazardType


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def fetch_inform_data(file):
    from ipc.models import Country

    # NOTE: set `data_only=True` to read cell value instead of formula
    workbook = openpyxl.load_workbook(file, data_only=True)
    worksheet = workbook.get_sheet_by_name('INFORM Risk 2022 (a-z)')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    # Iterate the loop to read the cell values
    data_all = []
    for i in range(4, max_rows + 2):
        country = worksheet.cell(row=i, column=1).value
        earthquake = worksheet.cell(row=i, column=9).value
        flood = worksheet.cell(row=i, column=10).value
        tsunami = worksheet.cell(row=i, column=11).value
        cyclone = worksheet.cell(row=i, column=12).value
        drought = worksheet.cell(row=i, column=13).value
        epidemic = worksheet.cell(row=i, column=14).value
        data = {
            'country': country,
            'earthquake': earthquake,
            'flood': flood,
            'tsunami': tsunami,
            'cyclone': cyclone,
            'drought': drought,
            'epidemic': epidemic
        }
        if Country.objects.filter(name=data['country']).exists():
            data_all.append(data.copy())
    for data in data_all:
        country = Country.objects.get(name=data['country'])
        earthquake_risk_score = data['earthquake']
        flood_risk_score = data['flood']
        tsunami_risk_score = data['tsunami']
        cyclone_risk_score = data['cyclone']
        drought_risk_score = data['drought']
        epidemic_risk_score = data['epidemic']
        data = {
            'country': country,
            'hazard_type': HazardType.EARTHQUAKE,
            'risk_score': earthquake_risk_score
        }
        InformRisk.objects.create(**data)
        data = {
            'country': country,
            'hazard_type': HazardType.FLOOD,
            'risk_score': flood_risk_score
        }
        InformRisk.objects.create(**data)
        data = {
            'country': country,
            'hazard_type': HazardType.CYCLONE,
            'risk_score': cyclone_risk_score
        }
        InformRisk.objects.create(**data)
        data = {
            'country': country,
            'hazard_type': HazardType.DROUGHT,
            'risk_score': drought_risk_score
        }
        InformRisk.objects.create(**data)
        data = {
            'country': country,
            'hazard_type': HazardType.EPIDEMIC,
            'risk_score': epidemic_risk_score
        }
        InformRisk.objects.create(**data)
        data = {
            'country': country,
            'hazard_type': HazardType.TSUNAMI,
            'risk_score': tsunami_risk_score
        }
        InformRisk.objects.create(**data)
