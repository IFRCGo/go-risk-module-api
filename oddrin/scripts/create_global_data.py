import openpyxl

from oddrin.models import (
    DisplacementData,
    HazardType
)


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_hazard_type(hazard_type):
    if hazard_type == 'Flood':
        hazard_type = HazardType.FLOOD
    elif hazard_type == 'Cyclone':
        hazard_type = HazardType.CYCLONE
    return hazard_type


def parse_empty_cell_value(data):
    if data == '':
        data = None
    return data


def create_global_displacment_data(file):
    from ipc.models import Country
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name('Displacement Risk Per Month')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    for i in range(2, max_rows):
        iso3 = worksheet.cell(row=i, column=2).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=3).value)
        annual_average_displacement = parse_empty_cell_value(worksheet.cell(row=i, column=4).value)
        january = parse_empty_cell_value(str(worksheet.cell(row=i, column=5).value).replace("%", ""))
        february = parse_empty_cell_value(str(worksheet.cell(row=i, column=6).value).replace("%", ""))
        march = parse_empty_cell_value(str(worksheet.cell(row=i, column=7).value).replace("%", ""))
        april = parse_empty_cell_value(str(worksheet.cell(row=i, column=8).value).replace("%", ""))
        may = parse_empty_cell_value(str(worksheet.cell(row=i, column=9).value).replace("%", ""))
        june = parse_empty_cell_value(str(worksheet.cell(row=i, column=10).value).replace("%", ""))
        july = parse_empty_cell_value(str(worksheet.cell(row=i, column=11).value).replace("%", ""))
        august = parse_empty_cell_value(str(worksheet.cell(row=i, column=12).value).replace("%", ""))
        september = parse_empty_cell_value(str(worksheet.cell(row=i, column=13).value).replace("%", ""))
        october = parse_empty_cell_value(str(worksheet.cell(row=i, column=14).value).replace("%", ""))
        november = parse_empty_cell_value(str(worksheet.cell(row=i, column=15).value).replace("%", ""))
        december = parse_empty_cell_value(str(worksheet.cell(row=i, column=16).value).replace("%", ""))
        if annual_average_displacement:
            if january:
                january = float(january) * annual_average_displacement
            if february:
                february = float(february) * annual_average_displacement
            if march:
                march = float(march) * annual_average_displacement
            if april:
                april = float(april) * annual_average_displacement
            if may:
                may = float(may) * annual_average_displacement
            if june:
                june = float(june) * annual_average_displacement
            if july:
                july = float(july) * annual_average_displacement
            if august:
                august = float(august) * annual_average_displacement
            if september:
                september = float(september) * annual_average_displacement
            if october:
                october = float(october) * annual_average_displacement
            if november:
                november = float(november) * annual_average_displacement
            if december:
                december = float(december) * annual_average_displacement

        if Country.objects.filter(iso3=iso3.lower()).exists():
            data = {
                'country': Country.objects.filter(iso3=iso3.lower()).first(),
                'iso3': iso3,
                'hazard_type': hazard_type,
                'january': january,
                'february': february,
                'march': march,
                'april': april,
                'may': may,
                'june': june,
                'july': july,
                'august': august,
                'september': september,
                'october': october,
                'november': november,
                'december': december,
                'annual_average_displacement': annual_average_displacement
            }
            DisplacementData.objects.get_or_create(**data)
