import openpyxl

from oddrin.models import IdmcSuddenOnset, Oddrin, HazardType


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def parse_hazard_type(hazard_type):
    if hazard_type == 'Earthquake':
        hazard_type = HazardType.EARTHQUAKE
    elif hazard_type == 'Storm-surge':
        hazard_type = HazardType.STORM
    elif hazard_type == 'Tsunami':
        hazard_type = HazardType.TSUNAMI
    elif hazard_type == 'Flood':
        hazard_type = HazardType.FLOOD
    elif hazard_type == 'Cyclonic Wind':
        hazard_type = HazardType.CYCLONE
    return hazard_type


def parse_empty_cell_value(data):
    if data == '':
        data = None
    return data


def fetch_idmc_pmd_data(file):
    from ipc.models import Country
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name('ADD-PMD')
    max_rows = get_maximum_rows(sheet_object=worksheet)
    for i in range(2, max_rows):
        country = worksheet.cell(row=i, column=2).value
        hazard_type = parse_hazard_type(worksheet.cell(row=i, column=3).value)
        annual_average_displacement = parse_empty_cell_value(worksheet.cell(row=i, column=4).value)
        return_period_20_years = parse_empty_cell_value(worksheet.cell(row=i, column=5).value)
        return_period_50_years = parse_empty_cell_value(worksheet.cell(row=i, column=6).value)
        return_period_100_years = parse_empty_cell_value(worksheet.cell(row=i, column=7).value)
        return_period_250_years = parse_empty_cell_value(worksheet.cell(row=i, column=8).value)
        return_period_1000_years = parse_empty_cell_value(worksheet.cell(row=i, column=9).value)
        return_period_1500_years = parse_empty_cell_value(worksheet.cell(row=i, column=10).value)
        if Country.objects.filter(name=country).exists():
            data = {
                'country': Country.objects.get(name=country),
                'hazard_type': hazard_type,
                'annual_average_displacement': annual_average_displacement,
                'return_period_20_years': return_period_20_years,
                'return_period_50_years': return_period_50_years,
                'return_period_100_years': return_period_100_years,
                'return_period_250_years': return_period_250_years,
                'return_period_1000_years': return_period_1000_years,
                'return_period_1500_years': return_period_1500_years
            }
            IdmcSuddenOnset.objects.create(**data)
